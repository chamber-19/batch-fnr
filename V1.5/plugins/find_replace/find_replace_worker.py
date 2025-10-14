"""
Find & Replace Worker for background processing.

Handles the actual find and replace operations in a separate thread
to keep the UI responsive during batch processing.
"""

import os
import time
import getpass
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

from PySide6.QtCore import QObject, Signal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from core.autocad_bridge import AutoCADBridge


class FindReplaceWorker(QObject):
    """
    Worker class for find and replace operations.
    
    Runs in a separate thread to process DWG files without blocking the UI.
    """
    
    # Signals
    finished = Signal(bool, str)  # success, message
    progress = Signal(int)  # progress percentage
    log_message = Signal(str)  # log messages
    file_processed = Signal(str, bool)  # file_path, success
    
    def __init__(self, settings: Dict[str, Any]):
        super().__init__()
        
        self.settings = settings
        self.cancelled = False
        
        # Processing data
        self.changes_log: List[Dict[str, Any]] = []
        self.error_log: List[Dict[str, Any]] = []
        
    def cancel(self):
        """Cancel the processing."""
        self.cancelled = True
    
    def process(self):
        """Main processing method."""
        try:
            self.log_message.emit("Starting find and replace processing...")
            
            # Get file list
            dwg_files = self._get_dwg_files()
            if not dwg_files:
                self.finished.emit(False, "No DWG files found to process")
                return
            
            self.log_message.emit(f"Found {len(dwg_files)} DWG files to process")
            
            # Get AutoCAD bridge
            bridge = self.settings.get('autocad_bridge')
            if not bridge:
                self.finished.emit(False, "AutoCAD bridge not available")
                return
            
            # Ensure AutoCAD connection
            try:
                bridge.ensure_connection()
            except Exception as e:
                self.finished.emit(False, f"Failed to connect to AutoCAD: {e}")
                return
            
            # Process files
            success_count = 0
            total_files = len(dwg_files)
            
            for i, file_path in enumerate(dwg_files):
                if self.cancelled:
                    self.log_message.emit("Processing cancelled by user")
                    break
                
                # Update progress
                progress = int((i / total_files) * 100)
                self.progress.emit(progress)
                
                # Process file
                try:
                    success = self._process_file(bridge, file_path)
                    if success:
                        success_count += 1
                    
                    self.file_processed.emit(file_path, success)
                    
                except Exception as e:
                    self.log_message.emit(f"Error processing {file_path}: {e}")
                    self.error_log.append({
                        "File": file_path,
                        "Error": str(e),
                        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })
                    self.file_processed.emit(file_path, False)
            
            # Final progress
            self.progress.emit(100)
            
            # Generate report
            if not self.cancelled:
                report_path = self._generate_report()
                
                if self.changes_log or self.error_log:
                    message = f"Processing complete. {success_count}/{total_files} files processed successfully."
                    if report_path:
                        message += f" Report saved to: {report_path}"
                else:
                    message = "Processing complete. No changes were made."
                
                self.finished.emit(True, message)
            else:
                self.finished.emit(False, "Processing was cancelled")
                
        except Exception as e:
            self.finished.emit(False, f"Processing failed: {e}")
    
    def _get_dwg_files(self) -> List[str]:
        """Get list of DWG files to process."""
        path = self.settings['path']
        mode = self.settings['mode']
        recursive = self.settings.get('recursive', False)
        
        dwg_files = []
        
        if mode == 'directory':
            # Directory mode
            if not os.path.isdir(path):
                return []
            
            path_obj = Path(path)
            
            if recursive:
                pattern = "**/*.dwg"
            else:
                pattern = "*.dwg"
            
            for dwg_file in path_obj.glob(pattern):
                if dwg_file.is_file():
                    dwg_files.append(str(dwg_file))
        
        else:
            # Files mode
            file_paths = path.split(';')
            for file_path in file_paths:
                file_path = file_path.strip()
                if file_path and os.path.isfile(file_path) and file_path.lower().endswith('.dwg'):
                    dwg_files.append(file_path)
        
        return sorted(dwg_files)
    
    def _process_file(self, bridge: AutoCADBridge, file_path: str) -> bool:
        """Process a single DWG file."""
        self.log_message.emit(f"Processing: {os.path.basename(file_path)}")
        
        try:
            # Open document
            doc = bridge.open_dwg(file_path)
            if not doc:
                raise RuntimeError("Failed to open document")
            
            file_changes = []
            changed = False
            
            # Apply stamp if configured
            stamp_settings = self.settings.get('stamp_settings', {})
            if stamp_settings.get('apply', False):
                issue_type = stamp_settings.get('issue_type', '')
                if issue_type and issue_type != '(leave as-is)':
                    if bridge.apply_stamp_layers(issue_type):
                        changed = True
                        file_changes.append({
                            "File": file_path,
                            "EntityType": "Stamp",
                            "OriginalText": "",
                            "NewText": f"Applied {issue_type} stamp",
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
            
            # Process text entities
            replacement_rules = self.settings.get('replacement_rules', [])
            if replacement_rules:
                text_changed = self._process_text_replacements(bridge, file_path, replacement_rules, file_changes)
                if text_changed:
                    changed = True
            
            # Process title block
            title_block_settings = self.settings.get('title_block_settings', {})
            if title_block_settings.get('enabled', False):
                tb_changed = self._process_title_block(bridge, file_path, title_block_settings, file_changes)
                if tb_changed:
                    changed = True
            
            # Save if changes were made and not in preview mode
            preview = self.settings.get('preview', False)
            if changed and not preview:
                bridge.save_doc()
                self.log_message.emit(f"Saved changes to: {os.path.basename(file_path)}")
            elif changed and preview:
                self.log_message.emit(f"Preview mode - changes not saved to: {os.path.basename(file_path)}")
            
            # Close document
            bridge.close_doc()
            
            # Add to changes log
            self.changes_log.extend(file_changes)
            
            return True
            
        except Exception as e:
            self.log_message.emit(f"Error processing {file_path}: {e}")
            
            # Try to close document if it was opened
            try:
                bridge.close_doc()
            except:
                pass
            
            raise
    
    def _process_text_replacements(self, bridge: AutoCADBridge, file_path: str, 
                                 replacement_rules: List[Dict[str, Any]], 
                                 file_changes: List[Dict[str, Any]]) -> bool:
        """Process text replacement rules."""
        changed = False
        
        # Collect text entities
        text_items = bridge.collect_text_entities()
        
        for text_item in text_items:
            original_text = text_item.text
            new_text = original_text
            
            # Apply all replacement rules
            for rule_data in replacement_rules:
                find_text = rule_data.get('find_text', '').strip()
                replace_text = rule_data.get('replace_text', '')
                ignore_case = rule_data.get('ignore_case', False)
                use_regex = rule_data.get('use_regex', False)
                
                if not find_text:
                    continue
                
                if use_regex:
                    # Regular expression replacement
                    import re
                    flags = re.IGNORECASE if ignore_case else 0
                    try:
                        new_text = re.sub(find_text, replace_text, new_text, flags=flags)
                    except re.error:
                        continue  # Skip invalid regex
                else:
                    # Simple text replacement
                    if ignore_case:
                        # Case-insensitive replacement
                        import re
                        pattern = re.escape(find_text)
                        new_text = re.sub(pattern, replace_text, new_text, flags=re.IGNORECASE)
                    else:
                        # Case-sensitive replacement
                        new_text = new_text.replace(find_text, replace_text)
            
            # Apply changes if text was modified
            if new_text != original_text:
                if not self.settings.get('preview', False):
                    bridge.set_entity_text(text_item.ent, new_text)
                
                file_changes.append({
                    "File": file_path,
                    "EntityType": text_item.entity_type,
                    "OriginalText": original_text,
                    "NewText": new_text,
                    "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "User": getpass.getuser()
                })
                
                changed = True
        
        return changed
    
    def _process_title_block(self, bridge: AutoCADBridge, file_path: str,
                           title_block_settings: Dict[str, Any],
                           file_changes: List[Dict[str, Any]]) -> bool:
        """Process title block operations."""
        changed = False
        
        # Get settings
        fields = title_block_settings.get('fields', {})
        shift_down = title_block_settings.get('shift_down', False)
        shift_up = title_block_settings.get('shift_up', False)
        clear_operations = title_block_settings.get('clear_operations', [])
        rev_section = title_block_settings.get('rev_section', 'Rev1')
        
        # Process block references in both spaces
        for space_name in ["ModelSpace", "PaperSpace"]:
            try:
                if space_name == "ModelSpace":
                    space = bridge.doc.ModelSpace
                else:
                    space = bridge.doc.PaperSpace
                
                for entity in space:
                    if hasattr(entity, 'EntityName') and entity.EntityName == "AcDbBlockReference":
                        if hasattr(entity, 'HasAttributes') and entity.HasAttributes:
                            # Process this block reference
                            if self._process_block_title_block(
                                bridge, entity, file_path, title_block_settings, file_changes
                            ):
                                changed = True
                                
            except Exception as e:
                self.log_message.emit(f"Error processing {space_name} title blocks: {e}")
        
        return changed
    
    def _process_block_title_block(self, bridge: AutoCADBridge, block_ref, file_path: str,
                                 title_block_settings: Dict[str, Any],
                                 file_changes: List[Dict[str, Any]]) -> bool:
        """Process title block operations on a single block reference."""
        changed = False
        
        # This is a simplified version - in a full implementation, you would
        # include all the title block processing logic from the original script
        
        def process_attribute(attr, tag, value):
            # Process individual attributes based on title block settings
            fields = title_block_settings.get('fields', {})
            
            # Check if this attribute should be updated
            if tag in fields:
                new_value = fields[tag]
                if new_value != value:
                    file_changes.append({
                        "File": file_path,
                        "EntityType": "Title Block Attribute",
                        "OriginalText": value,
                        "NewText": new_value,
                        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "User": getpass.getuser()
                    })
                    return new_value
            
            return None
        
        # Process attributes
        if bridge.process_block_attributes(block_ref, process_attribute):
            changed = True
        
        return changed
    
    def _generate_report(self) -> Optional[str]:
        """Generate Excel report of changes."""
        if not self.changes_log and not self.error_log:
            return None
        
        try:
            # Determine report directory
            path = self.settings['path']
            mode = self.settings['mode']
            
            if mode == 'directory':
                report_dir = path
            else:
                # Use directory of first file
                files = path.split(';')
                if files:
                    report_dir = os.path.dirname(files[0])
                else:
                    report_dir = os.getcwd()
            
            # Generate unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            preview_suffix = "_PREVIEW" if self.settings.get('preview', False) else ""
            report_filename = f"FindReplace_Report{preview_suffix}_{timestamp}.xlsx"
            report_path = os.path.join(report_dir, report_filename)
            
            # Create workbook
            wb = Workbook()
            
            # Changes sheet
            if self.changes_log:
                ws_changes = wb.active
                ws_changes.title = "Changes"
                
                # Headers
                headers = ["File", "EntityType", "OriginalText", "NewText", "Timestamp", "User"]
                for col, header in enumerate(headers, 1):
                    cell = ws_changes.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                # Data
                for row, change in enumerate(self.changes_log, 2):
                    for col, header in enumerate(headers, 1):
                        ws_changes.cell(row=row, column=col, value=change.get(header, ""))
            
            # Errors sheet
            if self.error_log:
                ws_errors = wb.create_sheet("Errors")
                
                # Headers
                headers = ["File", "Error", "Timestamp"]
                for col, header in enumerate(headers, 1):
                    cell = ws_errors.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                # Data
                for row, error in enumerate(self.error_log, 2):
                    for col, header in enumerate(headers, 1):
                        ws_errors.cell(row=row, column=col, value=error.get(header, ""))
            
            # Save workbook
            wb.save(report_path)
            
            self.log_message.emit(f"Report saved: {report_path}")
            return report_path
            
        except Exception as e:
            self.log_message.emit(f"Failed to generate report: {e}")
            return None
