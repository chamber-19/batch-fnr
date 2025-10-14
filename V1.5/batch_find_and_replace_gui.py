import sys
import os
import getpass
import re
import time
import pythoncom
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from PIL import ImageTk, Image
import win32com.client
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import logging
import threading
import queue
import smtplib
from email.utils import formataddr
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
import pywintypes  # COM errors (E_FAIL)

def resource_path(rel):
    base = getattr(sys, '_MEIPASS', os.path.dirname(__file__))
    return os.path.join(base, rel)

LOGO_PATH = resource_path("root3_logo.png")


# Email configuration (user provided)
SENDER_EMAIL = "Hyphaeos@gmail.com"
APP_PASSWORD = "ibhb mbcg wluz byri"
RECEIVER_EMAIL = "dustin.ward@root3power.com"

# UI constants
COLOR_BG = '#121212'
COLOR_FG = '#1E90FF'
COLOR_ENTRY_BG = '#2C2C2C'
COLOR_ENTRY_FG = '#FFFFFF'
COLOR_BUTTON = '#000080'
FONT_HEADER = ("Arial", 14, "bold")
FONT_LABEL = ("Arial", 10, "bold")
FONT_ENTRY = ("Arial", 11)
MAX_REPLACEMENTS = 10
ENTRY_WIDTH = 40
PROGRESS_LENGTH = 850
LOGO_FILE = "root3_logo.png"
VERSION = "1.2"

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------- small helpers ----------
def today_mm_dd_yy() -> str:
    return datetime.now().strftime("%m-%d-%y")

def ensure_unique_path(path: str) -> str:
    base, ext = os.path.splitext(path)
    n, cand = 1, path
    while os.path.exists(cand):
        cand = f"{base} ({n}){ext}"
        n += 1
    return cand

def compute_report_dir(mode: str, directory: str | None, files: list[str] | None) -> str:
    if mode == "directory":
        if not directory or not os.path.isdir(directory):
            raise NotADirectoryError(directory or "(missing)")
        return os.path.abspath(directory)
    files = [os.path.abspath(f) for f in (files or []) if f]
    if not files:
        raise ValueError("No files selected")
    try:
        common = os.path.commonpath(files)
    except ValueError:
        common = os.path.dirname(files[0])
    return common if os.path.isdir(common) else os.path.dirname(common)

def open_in_default_app(path: str) -> None:
    try:
        os.startfile(path)  # Windows
    except Exception as e:
        logger.error(f"Auto-open failed: {e}")

def _is_e_fail(err: Exception) -> bool:
    """Why: E_FAIL (-2147467259) is noisy; treat as benign."""
    return isinstance(err, pywintypes.com_error) and getattr(err, "hresult", None) == -2147467259

class AutoCADReplacerGUI:
    """GUI for batch find/replace in DWGs with Title Block & Stamp helpers."""

    def __init__(self, root):
        self.root = root
        self.root.title("Batch Find & Replace")
        WIDTH, HEIGHT = 1440, 840
        self.root.geometry(f"{WIDTH}x{HEIGHT}")
        self.root.minsize(1200, 720)
        try:
            self.root.update_idletasks()
            sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
            x, y = (sw - WIDTH) // 2, (sh - HEIGHT) // 2
            self.root.geometry(f"{WIDTH}x{HEIGHT}+{x}+{y}")
        except Exception:
            pass
        self.root.configure(bg=COLOR_BG)
        self.root.resizable(True, True)

        # Icon/logo
        self.logo = None
        try:
            self.logo = ImageTk.PhotoImage(Image.open(resource_path(LOGO_FILE)).resize((80, 80)))
            self.root.iconphoto(True, self.logo)
        except Exception as e:
            logger.error(f"Could not load {LOGO_FILE}: {e}")
            messagebox.showwarning("Logo Load Error", f"Could not load logo: {e}. Continuing without logo.")

        # Header
        header_frame = tk.Frame(root, bg=COLOR_BG)
        header_frame.grid(row=0, column=0, pady=(5, 5), padx=10, sticky='nsew', columnspan=2)
        root.grid_rowconfigure(0, weight=0)
        root.grid_columnconfigure(0, weight=1)
        if self.logo:
            tk.Label(header_frame, image=self.logo, bg=COLOR_BG).pack(side=tk.LEFT, padx=(0, 10))
        ht = tk.Frame(header_frame, bg=COLOR_BG); ht.pack(side=tk.LEFT, expand=True)
        tk.Label(ht, text="Root3Power LLC", fg=COLOR_FG, bg=COLOR_BG, font=FONT_HEADER).pack(pady=(0, 5))
        tk.Label(ht, text="Batch Find & Replace", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack()

        # Layout rows
        for r in range(1, 8):
            root.grid_rowconfigure(r, weight=0)
        root.grid_rowconfigure(6, weight=1)
        root.grid_columnconfigure(0, weight=1)

        # Project Selection
        pf = tk.Frame(root, bg=COLOR_BG, padx=10, pady=5); pf.grid(row=1, column=0, sticky='nsew')
        tk.Label(pf, text="Project Selection", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(anchor='w')
        self.mode_var = tk.StringVar(value="directory")
        mf = tk.Frame(pf, bg=COLOR_BG); mf.pack(pady=5, fill='x')
        tk.Radiobutton(mf, text="Entire Project Directory (Recursive)", variable=self.mode_var, value="directory",
                       command=self.update_mode, takefocus=False, bg=COLOR_BG, fg=COLOR_ENTRY_FG,
                       selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       indicatoron=False, font=FONT_LABEL).pack(side=tk.LEFT, padx=5)
        tk.Radiobutton(mf, text="Select Specific Files", variable=self.mode_var, value="files",
                       command=self.update_mode, takefocus=False, bg=COLOR_BG, fg=COLOR_ENTRY_FG,
                       selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       indicatoron=False, font=FONT_LABEL).pack(side=tk.LEFT, padx=5)
        pth = tk.Frame(pf, bg=COLOR_BG); pth.pack(pady=5, fill='x')
        self.path_entry = tk.Entry(pth, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', font=FONT_ENTRY)
        self.path_entry.pack(side=tk.LEFT, padx=5, fill='x', expand=True)
        tk.Button(pth, text="Browse", command=self.browse_path, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                  font=FONT_ENTRY, padx=10, pady=5, takefocus=False, relief="flat").pack(side=tk.LEFT, padx=5)

        # Rules
        rf = tk.Frame(root, bg=COLOR_BG, padx=10, pady=5); rf.grid(row=2, column=0, sticky='nsew')
        tk.Label(rf, text="Text Replacement Rules", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(anchor='w')
        lf = tk.Frame(rf, bg=COLOR_BG); lf.pack(fill='x')
        tk.Label(lf, text="Find", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).pack(side=tk.LEFT, padx=(0, 135))
        tk.Label(lf, text="Replace With", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).pack(side=tk.LEFT, padx=(110, 40))
        self.replacements_frame = tk.Frame(rf, bg=COLOR_BG); self.replacements_frame.pack(fill='x')
        self.replacements = []; self.add_replacement_pair()
        tk.Button(rf, text="Add Rule", command=self.add_replacement_pair, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                  font=FONT_ENTRY, padx=10, pady=5, takefocus=False, relief="flat").pack(pady=5)

        # Options
        of = tk.Frame(root, bg=COLOR_BG, padx=10, pady=5); of.grid(row=3, column=0, sticky='nsew')
        tk.Label(of, text="Options", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(anchor='w')
        self.preview_var = tk.BooleanVar(value=False)
        tk.Checkbutton(of, text="Preview Mode (No Changes Saved)", variable=self.preview_var,
                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON,
                       activeforeground=COLOR_ENTRY_FG, font=FONT_ENTRY, takefocus=False, indicatoron=False).pack(side=tk.LEFT, padx=5)
        self.recursive_var = tk.BooleanVar(value=False)
        tk.Checkbutton(of, text="Include Subdirectories", variable=self.recursive_var,
                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON,
                       activeforeground=COLOR_ENTRY_FG, font=FONT_ENTRY, takefocus=False, indicatoron=False).pack(side=tk.LEFT, padx=5)

        # Title Block
        tb = tk.Frame(root, bg=COLOR_BG, padx=10, pady=6); tb.grid(row=4, column=0, sticky='nsew')
        tk.Label(tb, text="Title Block", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=0, column=0, columnspan=16, sticky='w')
        self.tb_enable = tk.BooleanVar(value=False)
        tk.Checkbutton(tb, text="Enable", variable=self.tb_enable, fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON,
                       activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG, font=FONT_ENTRY, takefocus=False,
                       indicatoron=False).grid(row=0, column=16, sticky='e', padx=6)

        tk.Label(tb, text="Rev Section", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=1, column=0, sticky='e')
        self.tb_rev_choice = tk.StringVar(value="Rev1")
        ttk.Combobox(tb, textvariable=self.tb_rev_choice, state="readonly", width=8,
                     values=["Rev1","Rev2","Rev3","Rev4","Rev5"]).grid(row=1, column=1, padx=4, sticky='w')
        tk.Label(tb, text="REV", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=1, column=2, sticky='e')
        self.tb_rev_val = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_rev_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                 insertbackground='white', width=8).grid(row=1, column=3, padx=4, sticky='w')
        tk.Label(tb, text="DESC", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=1, column=4, sticky='e')
        self.tb_desc_val = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_desc_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                 insertbackground='white', width=36).grid(row=1, column=5, padx=4, sticky='w')
        tk.Label(tb, text="BY", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=1, column=6, sticky='e')
        self.tb_by_val = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_by_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                 insertbackground='white', width=10).grid(row=1, column=7, padx=4, sticky='w')
        tk.Label(tb, text="CHK", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=1, column=8, sticky='e')
        self.tb_chk_val = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_chk_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                 insertbackground='white', width=10).grid(row=1, column=9, padx=4, sticky='w')
        tk.Label(tb, text="DATE", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=1, column=10, sticky='e')
        self.tb_date_val = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_date_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                 insertbackground='white', width=12).grid(row=1, column=11, padx=4, sticky='w')

        row2 = 2
        tk.Label(tb, text="DWNBY", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=row2, column=0, sticky='e', pady=(6,0))
        self.tb_dwnby = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_dwnby, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', width=12).grid(row=row2, column=1, padx=4, pady=(6,0), sticky='w')
        tk.Label(tb, text="DWNDATE", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=row2, column=2, sticky='e', pady=(6,0))
        self.tb_dwndate = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_dwndate, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', width=12).grid(row=row2, column=3, padx=4, pady=(6,0), sticky='w')
        tk.Label(tb, text="CHKBY", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=row2, column=4, sticky='e', pady=(6,0))
        self.tb_chkby = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_chkby, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', width=12).grid(row=row2, column=5, padx=4, pady=(6,0), sticky='w')
        tk.Label(tb, text="CHKDATE", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=row2, column=6, sticky='e', pady=(6,0))
        self.tb_chkdate = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_chkdate, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', width=12).grid(row=row2, column=7, padx=4, pady=(6,0), sticky='w')
        tk.Label(tb, text="ENGR", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=row2, column=8, sticky='e', pady=(6,0))
        self.tb_engr = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_engr, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', width=12).grid(row=row2, column=9, padx=4, pady=(6,0), sticky='w')
        tk.Label(tb, text="ENGRDATE", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=row2, column=10, sticky='e', pady=(6,0))
        self.tb_engrdate = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_engrdate, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', width=12).grid(row=row2, column=11, padx=4, pady=(6,0), sticky='w')
        tk.Label(tb, text="Revision (REV)", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=row2, column=12, sticky='e', pady=(6,0))
        self.tb_overall_rev = tk.StringVar()
        tk.Entry(tb, textvariable=self.tb_overall_rev, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', width=8).grid(row=row2, column=13, padx=4, pady=(6,0), sticky='w')

        # Shift / Clear
        row3 = 3
        self.tb_shift_down = tk.BooleanVar(value=False)
        self.tb_shift_up = tk.BooleanVar(value=False)
        tk.Checkbutton(tb, text="Shift Revisions Down (1→2→…→5)", variable=self.tb_shift_down,
                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON,
                       activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       font=FONT_ENTRY, takefocus=False, indicatoron=False).grid(row=row3, column=0, columnspan=4, sticky='w', pady=(10,0))
        tk.Checkbutton(tb, text="Shift Revisions Up (5→4→…→1)", variable=self.tb_shift_up,
                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON,
                       activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       font=FONT_ENTRY, takefocus=False, indicatoron=False).grid(row=row3, column=4, columnspan=4, sticky='w', pady=(10,0))

        clear_frame = tk.Frame(tb, bg=COLOR_BG); clear_frame.grid(row=4, column=0, columnspan=12, sticky='w', pady=(10, 0))
        tk.Label(clear_frame, text="Clear Rev Lines:", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=0, column=0, sticky='w')
        self.tb_clear_choice = tk.StringVar(value="Rev")
        ttk.Combobox(clear_frame, textvariable=self.tb_clear_choice, state="readonly", width=8,
                     values=["Rev","Rev2","Rev3","Rev4","Rev5"]).grid(row=0, column=1, padx=6, sticky='w')
        tk.Button(clear_frame, text="Add Clear-Rev", command=self._ui_add_clear_rev,
                  bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG, relief="flat").grid(row=0, column=2, padx=6)
        tk.Button(clear_frame, text="Remove Selected", command=self._ui_remove_selected_clear_rev,
                  bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG, relief="flat").grid(row=0, column=3, padx=6)
        self.tb_clear_list = tk.Listbox(clear_frame, height=3, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG)
        self.tb_clear_list.grid(row=1, column=0, columnspan=6, sticky='we', pady=(6,0))
        clear_frame.grid_columnconfigure(5, weight=1)

        # Stamp
        stamp = tk.Frame(tb, bg=COLOR_BG); stamp.grid(row=5, column=0, columnspan=12, sticky='w', pady=(10, 0))
        tk.Label(stamp, text="Stamp", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=0, column=0, sticky='w')
        tk.Label(stamp, text="Issue:", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).grid(row=0, column=1, padx=(10,2), sticky='e')
        self.stamp_issue = tk.StringVar(value="(leave as-is)")
        ttk.Combobox(stamp, textvariable=self.stamp_issue, state="readonly", width=20,
                     values=["(leave as-is)", "APPROVAL", "PRELIM", "CONSTRUCTION", "BID", "AS-BUILT", "REFERENCE"]).grid(row=0, column=2, padx=6, sticky='w')
        self.stamp_apply = tk.BooleanVar(value=True)
        tk.Checkbutton(stamp, text="Apply in run", variable=self.stamp_apply,
                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON,
                       activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       font=FONT_ENTRY, takefocus=False, indicatoron=False).grid(row=0, column=3, padx=10, sticky='w')

        # Run + Progress
        run = tk.Frame(root, bg=COLOR_BG, padx=10, pady=5); run.grid(row=5, column=0, sticky='nsew')
        self.run_btn = tk.Button(run, text="Run", command=self.start_process, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                                 font=FONT_ENTRY, padx=10, pady=5, takefocus=False, relief="flat"); self.run_btn.pack(side=tk.LEFT, padx=5)
        self.cancel_btn = tk.Button(run, text="Cancel", command=self.cancel_process, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                                    font=FONT_ENTRY, padx=10, pady=5, state=tk.DISABLED, takefocus=False, relief="flat"); self.cancel_btn.pack(side=tk.LEFT, padx=5)
        self.progress = ttk.Progressbar(run, length=PROGRESS_LENGTH, mode='determinate', style='black.Horizontal.TProgressbar')
        self.progress.pack(pady=5, fill='x')

        # --- TERMINAL LOG (scrolled) ---
        log_wrap = tk.Frame(root, bg=COLOR_BG)
        log_wrap.grid(row=6, column=0, padx=10, pady=(6, 10), sticky='nsew')
        tk.Label(log_wrap, text="Terminal Log", bg=COLOR_BG, fg=COLOR_FG, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.status_log = scrolledtext.ScrolledText(
            log_wrap, height=10, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
            insertbackground=COLOR_ENTRY_FG, font=FONT_ENTRY, state='disabled'
        )
        self.status_log.pack(fill="both", expand=True)

        # Bottom
        bottom = tk.Frame(root, bg=COLOR_BG); bottom.grid(row=7, column=0, pady=(0, 5), sticky='se')
        tk.Button(bottom, text="Suggest", command=self.open_suggestion_window, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                  font=FONT_ENTRY, relief="flat").pack(side="left", padx=5)
        tk.Button(bottom, text="Report Bug", command=self.open_bug_report_window, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                  font=FONT_ENTRY, relief="flat").pack(side="left", padx=5)
        tk.Label(bottom, text=f"Version {VERSION}", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=("Arial", 8)).pack(side="right")

        self.log_message("Ready")

        style = ttk.Style()
        style.configure('black.Horizontal.TProgressbar', background='blue', troughcolor='gray',
                        bordercolor='white', lightcolor='white', darkcolor='white')

        # Threading
        self.process_thread = None
        self.cancel_event = threading.Event()
        self.queue = queue.Queue()

        # Cache
        self.blockdef_cache: dict[str, object | None] = {}

    # ---------- UI helpers ----------
    def create_tooltip(self, widget, text):
        def enter(event):
            self.tooltip = tk.Toplevel(widget)
            self.tooltip.wm_overrideredirect(True)
            self.tooltip.wm_geometry(f"+{event.x_root + 20}+{event.y_root + 20}")
            tk.Label(self.tooltip, text=text, bg='yellow', fg='black', relief='solid', borderwidth=1, padx=5, pady=3).pack()
        def leave(event):
            if hasattr(self, 'tooltip'):
                self.tooltip.destroy()
        widget.bind("<Enter>", enter); widget.bind("<Leave>", leave)

    def log_message(self, message, level='info'):
        self.status_log.config(state='normal')
        self.status_log.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.status_log.see(tk.END)
        self.status_log.config(state='disabled')
        (logger.error if level == 'error' else logger.info)(message)

    def update_mode(self):
        self.path_entry.delete(0, tk.END)

    def browse_path(self):
        if self.mode_var.get() == "directory":
            path = filedialog.askdirectory()
            if path:
                self.path_entry.delete(0, tk.END); self.path_entry.insert(0, path)
        else:
            paths = filedialog.askopenfilenames(filetypes=[("DWG Files", "*.dwg")])
            if paths:
                self.path_entry.delete(0, tk.END); self.path_entry.insert(0, ";".join(paths))

    # ---------- Rules UI ----------
    def add_replacement_pair(self):
        if len(self.replacements) >= MAX_REPLACEMENTS:
            self.log_message("Error: Maximum of 10 find-replace pairs allowed.", 'error')
            return
        frame = tk.Frame(self.replacements_frame, bg=COLOR_BG); frame.pack(pady=2, fill='x')
        find_entry = tk.Entry(frame, width=ENTRY_WIDTH, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', font=FONT_ENTRY)
        find_entry.pack(side=tk.LEFT, padx=5, fill='x', expand=True)
        tk.Label(frame, text="->", fg=COLOR_ENTRY_FG, bg=COLOR_BG).pack(side=tk.LEFT)
        replace_entry = tk.Entry(frame, width=ENTRY_WIDTH, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', font=FONT_ENTRY)
        replace_entry.pack(side=tk.LEFT, padx=5, fill='x', expand=True)
        case_var = tk.BooleanVar()
        tk.Checkbutton(frame, text="Ignore Case", variable=case_var, fg=COLOR_ENTRY_FG, bg=COLOR_BG,
                       selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       font=FONT_ENTRY, takefocus=False, indicatoron=False).pack(side=tk.LEFT, padx=2)
        regex_var = tk.BooleanVar()
        tk.Checkbutton(frame, text="Use Advanced Patterns", variable=regex_var, fg=COLOR_ENTRY_FG, bg=COLOR_BG,
                       selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       font=FONT_ENTRY, takefocus=False, indicatoron=False).pack(side=tk.LEFT, padx=2)
        tk.Button(frame, text="Remove", command=lambda f=frame: self.remove_pair(f), bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                  font=FONT_ENTRY, padx=10, pady=5, takefocus=False, relief="flat").pack(side=tk.LEFT)
        self.replacements.append((find_entry, replace_entry, case_var, regex_var))

    def remove_pair(self, frame):
        frame.destroy()
        self.replacements = [pair for pair in self.replacements if pair[0].winfo_exists()]

    def validate_replacements(self, replacements):
        valid = []
        for find_entry, _, case_var, regex_var in replacements:
            find = find_entry.get().strip()
            if not find:
                continue
            if regex_var.get():
                try:
                    re.compile(find)
                except re.error as e:
                    self.log_message(f"Invalid regex '{find}': {e}", 'error')
                    return None
            valid.append((find_entry, _ , case_var, regex_var))
        return valid

    # ---------- AutoCAD helpers ----------
    def get_autocad_app(self):
        try:
            acad = win32com.client.Dispatch("AutoCAD.Application")
            acad.Visible = True
            return acad
        except Exception as e:
            self.log_message(f"Error connecting to AutoCAD: {e}", 'error')
            return None

    def process_entity_text(self, entity, entity_type, original_text, replacements, preview, changes_list, file_path):
        new_text = original_text
        for find_entry, replace_entry, case_var, regex_var in replacements:
            find = find_entry.get().strip()
            replace = replace_entry.get().strip()
            if not find:
                continue
            if regex_var.get():
                flags = re.IGNORECASE if case_var.get() else 0
                new_text = re.sub(find, replace, new_text, flags=flags)
            elif case_var.get():
                lower_text = new_text.lower()
                new_text = lower_text.replace(find.lower(), replace)
                if new_text != original_text.lower():
                    new_text = original_text.replace(original_text.lower().replace(find.lower(), replace), new_text)
            else:
                new_text = new_text.replace(find, replace)

        if new_text != original_text:
            if not preview:
                try:
                    if entity_type == "AcDbText":
                        entity.TextString = new_text
                    elif entity_type == "AcDbMText":
                        try:
                            entity.Text = new_text
                        except Exception:
                            entity.Contents = new_text
                    elif entity_type == "AcDbMLeader" and hasattr(entity, "MTextContent"):
                        entity.MTextContent = new_text
                    elif entity_type.startswith("AcDbDimension"):
                        entity.TextOverride = new_text
                except Exception as e:
                    self.log_message(f"Warning: failed to write text on {file_path}: {e}", "error")
            changes_list.append({
                "File": file_path,
                "EntityType": entity_type,
                "OriginalText": original_text,
                "NewText": new_text,
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "User": getpass.getuser()
            })
            return True
        return False

    def _read_attr_map(self, entity):
        out = {}
        if getattr(entity, "HasAttributes", False):
            for attr in entity.GetAttributes():
                tag = getattr(attr, "TagString", "")
                if tag:
                    out[tag.upper().strip()] = attr.TextString
        return out

    def _write_attr_if_exists(self, entity, tag, value):
        wrote = False
        if getattr(entity, "HasAttributes", False):
            for attr in entity.GetAttributes():
                if getattr(attr, "TagString", "").upper().strip() == tag.upper():
                    if attr.TextString != value:
                        attr.TextString = value
                    wrote = True
        return wrote

    def _clear_rev_line_on_blockref(self, entity, n, preview, changes_list, file_path):
        changed = False
        try:
            amap = self._read_attr_map(entity)
            if not amap:
                return False
            for base in ["REV", "DESC", "BY", "DATE", "CHK", "CHKBY"]:
                tag = f"{base}{n}"
                if tag in amap:
                    if not preview:
                        self._write_attr_if_exists(entity, tag, "")
                    changes_list.append({
                        "File": file_path,
                        "EntityType": f"ClearRev {n}",
                        "OriginalText": amap.get(tag, ""),
                        "NewText": "",
                        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "User": getpass.getuser()
                    })
                    changed = True
        except Exception as e:
            self.log_message(f"Clear Rev error: {e}", "error")
        return changed

    def _shift_revisions_on_blockref(self, entity, direction, preview, changes_list, file_path):
        if direction not in ("down", "up"): return False
        changed = False
        try:
            amap = self._read_attr_map(entity)
            if not amap:
                return False
            bases = ["REV", "DESC", "BY", "DATE", "CHK", "CHKBY"]
            indices = list(range(1, 6))
            if direction == "down":
                src_order, dest_shift = indices[:-1], +1
            else:
                src_order, dest_shift = indices[:0:-1], -1
            planned = {}
            for n in src_order:
                dest = n + dest_shift
                for base in bases:
                    src_tag, dest_tag = f"{base}{n}", f"{base}{dest}"
                    if src_tag in amap and dest_tag in amap:
                        planned[dest_tag] = amap[src_tag]
            if planned:
                for tag, val in planned.items():
                    try:
                        if not preview:
                            self._write_attr_if_exists(entity, tag, val)
                        changes_list.append({
                            "File": file_path, "EntityType": f"ShiftRevisions {direction}",
                            "OriginalText": amap.get(tag, ""), "NewText": val,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
                    except Exception as e:
                        self.log_message(f"Shift write failed for {tag}: {e}", "error")
        except Exception as e:
            self.log_message(f"Shift error: {e}", "error")
        return changed

    def process_block_attributes(self, entity, replacements, preview, changes_list, file_path):
        changed = False
        try:
            if entity.HasAttributes:
                for attr in entity.GetAttributes():
                    original_text = attr.TextString
                    new_text = original_text
                    for find_entry, replace_entry, case_var, regex_var in replacements:
                        find = find_entry.get().strip()
                        replace = replace_entry.get().strip()
                        if not find:
                            continue
                        if regex_var.get():
                            flags = re.IGNORECASE if case_var.get() else 0
                            new_text = re.sub(find, replace, new_text, flags=flags)
                        elif case_var.get():
                            lower_text = new_text.lower()
                            new_text = lower_text.replace(find.lower(), replace)
                            if new_text != original_text.lower():
                                new_text = original_text.replace(original_text.lower().replace(find.lower(), replace), new_text)
                        else:
                            new_text = new_text.replace(find, replace)
                    if new_text != original_text:
                        if not preview:
                            try:
                                attr.TextString = new_text
                            except Exception as e:
                                self.log_message(f"Warning: failed to write attribute on {file_path}: {e}", "error")
                                continue
                        changes_list.append({
                            "File": file_path, "EntityType": "AcDbBlockReference Attribute",
                            "OriginalText": original_text, "NewText": new_text,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
        except Exception as e:
            self.log_message(f"Error accessing attributes: {e}", "error")
        return changed

    def apply_title_block_to_blockref(self, entity, preview, changes_list, file_path):
        if not self.tb_enable.get():
            return False
        if not getattr(entity, "HasAttributes", False):
            return False

        choice = (self.tb_rev_choice.get() or "Rev1").strip().lower()
        try:
            n = int(choice.replace("rev", "")); n = max(1, min(5, n))
        except Exception:
            n = 1

        targets: dict[str, str] = {}
        def set_target(tag: str, value: str | None):
            v = (value or "").strip()
            if v: targets[tag] = v

        set_target(f"REV{n}",  self.tb_rev_val.get())
        set_target(f"DESC{n}", self.tb_desc_val.get())
        set_target(f"BY{n}",   self.tb_by_val.get())
        chk_val = (self.tb_chk_val.get() or "").strip()
        if chk_val:
            set_target(f"CHK{n}", chk_val)
            set_target(f"CHKBY{n}", chk_val)
        set_target(f"DATE{n}", self.tb_date_val.get())
        set_target("DWNBY",    self.tb_dwnby.get())
        set_target("DWNDATE",  self.tb_dwndate.get())
        set_target("CHKBY",    self.tb_chkby.get())
        set_target("CHKDATE",  self.tb_chkdate.get())
        set_target("ENGR",     self.tb_engr.get())
        set_target("ENGRDATE", self.tb_engrdate.get())
        set_target("REV",      self.tb_overall_rev.get())

        if not targets: return False

        changed = False
        try:
            for attr in entity.GetAttributes():
                tag = (getattr(attr, "TagString", "") or "").upper().strip()
                if not tag: continue
                if tag in targets:
                    new_val, old_val = targets[tag], attr.TextString
                    if new_val != old_val:
                        if not preview:
                            try: attr.TextString = new_val
                            except Exception as e:
                                self.log_message(f"Title Block write failed for {tag}: {e}", "error"); continue
                        changes_list.append({
                            "File": file_path, "EntityType": "TitleBlock Attribute",
                            "OriginalText": old_val, "NewText": new_val,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
                elif (tag.startswith("CHK") and tag.endswith(str(n))) and chk_val:
                    old_val = attr.TextString
                    if old_val != chk_val:
                        if not preview:
                            try: attr.TextString = chk_val
                            except Exception as e:
                                self.log_message(f"Title Block write failed for {tag}: {e}", "error"); continue
                        changes_list.append({
                            "File": file_path, "EntityType": "TitleBlock Attribute",
                            "OriginalText": old_val, "NewText": chk_val,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
        except Exception as e:
            self.log_message(f"Title Block error: {e}", "error")
        return changed

    @staticmethod
    def _is_xref_block_name(name: str) -> bool:
        if not name: return False
        u = str(name).upper()
        return ('|' in u) or u.startswith('*X')

    def _safe_get_block_def(self, doc, block_name: str):
        key = block_name or ""
        if key in self.blockdef_cache:
            return self.blockdef_cache[key]
        if self._is_xref_block_name(key):
            self.blockdef_cache[key] = None
            return None
        try:
            blk = doc.Blocks.Item(block_name)
        except Exception:
            self.blockdef_cache[key] = None
            return None
        # If the block def itself is xref-derived, skip scanning its definition (attributes still handled on refs).
        for flag in ("IsXRef", "IsXRefDependent", "IsFromExternalReference"):
            try:
                if bool(getattr(blk, flag)):
                    self.blockdef_cache[key] = None
                    return None
            except Exception:
                pass
        self.blockdef_cache[key] = blk
        return blk

    def process_block_reference(self, entity, replacements, preview, changes_list, file_path, *, depth: int = 0, max_depth: int = 2):
        changed = False

        # 0) Clear lines
        for item in list(self.tb_clear_list.get(0, tk.END)):
            choice = item.strip().lower()
            n = 1 if choice == "rev" else int(choice.replace("rev", ""))
            if self._clear_rev_line_on_blockref(entity, n, preview, changes_list, file_path):
                changed = True

        # 1) Shift
        shift_dir = "down" if self.tb_shift_down.get() else ("up" if self.tb_shift_up.get() else None)
        if shift_dir:
            try:
                if self._shift_revisions_on_blockref(entity, shift_dir, preview, changes_list, file_path):
                    changed = True
            except Exception as e:
                self.log_message(f"Error shifting revisions: {e}", "error")

        # 2) Attr replace
        if self.process_block_attributes(entity, replacements, preview, changes_list, file_path):
            changed = True

        # 3) Explicit TB writes
        if self.apply_title_block_to_blockref(entity, preview, changes_list, file_path):
            changed = True

        # 4) Nested scan (keep on; silence E_FAIL noise)
        if depth >= max_depth:
            return changed
        try:
            if getattr(entity, "IsXRef", False):
                return changed

            block_name = getattr(entity, "EffectiveName", None) or getattr(entity, "Name", None)
            if not block_name:
                return changed
            if self._is_xref_block_name(block_name):
                return changed

            block_def = self._safe_get_block_def(entity.Document, block_name)
            if block_def is None:
                return changed

            for ent in block_def:
                try:
                    ent_type = ent.EntityName
                except pywintypes.com_error as e:
                    if _is_e_fail(e):
                        continue  # silent E_FAIL on reading type
                    self.log_message(f"Nested entity type error in {block_name}: {e}", "error")
                    continue
                except Exception:
                    continue

                original_text = None
                try:
                    if ent_type == "AcDbText":
                        original_text = ent.TextString
                    elif ent_type == "AcDbMText":
                        try: original_text = ent.Text
                        except Exception: original_text = getattr(ent, "Contents", "")
                    elif ent_type.startswith("AcDbDimension"):
                        original_text = getattr(ent, "TextOverride", "") or ""
                    elif ent_type == "AcDbMLeader" and hasattr(ent, "MTextContent"):
                        original_text = ent.MTextContent
                except pywintypes.com_error as e:
                    if _is_e_fail(e):
                        continue  # silent E_FAIL on reading text
                    self.log_message(f"Nested read error in {block_name}: {e}", "error")
                    continue
                except Exception:
                    original_text = None

                if original_text is not None:
                    if self.process_entity_text(ent, ent_type, original_text, replacements, preview, changes_list, file_path):
                        changed = True
                    continue

                if ent_type == "AcDbBlockReference":
                    try:
                        if self.process_block_reference(ent, replacements, preview, changes_list, file_path,
                                                        depth=depth+1, max_depth=max_depth):
                            changed = True
                    except pywintypes.com_error as e:
                        if _is_e_fail(e):
                            continue  # silent
                        self.log_message(f"Nested recurse error in {block_name}: {e}", "error")
                    except Exception as e:
                        self.log_message(f"Nested recurse error in {block_name}: {e}", "error")
                    continue

                if ent_type == "AcDbTable":
                    try:
                        if self.process_table(ent, replacements, preview, changes_list, file_path):
                            changed = True
                    except pywintypes.com_error as e:
                        if _is_e_fail(e):
                            continue
                        self.log_message(f"Nested table error in {block_name}: {e}", "error")
                    except Exception as e:
                        self.log_message(f"Nested table error in {block_name}: {e}", "error")
                    continue

        except pywintypes.com_error as e:
            if _is_e_fail(e):
                pass  # full-block E_FAIL during nested scan: silent
            else:
                self.log_message(f"Nested scan COM error in {getattr(entity, 'EffectiveName', 'Unknown')}: {e}", "error")
        except Exception as e:
            self.log_message(f"Nested scan error in {getattr(entity, 'EffectiveName', 'Unknown')}: {e}", "error")

        return changed

    def process_table(self, entity, replacements, preview, changes_list, file_path):
        changed = False
        for row in range(entity.Rows):
            for col in range(entity.Columns):
                original_text = entity.GetCellValue(row, col)
                if isinstance(original_text, str):
                    new_text = original_text
                    for find_entry, replace_entry, case_var, regex_var in replacements:
                        find = find_entry.get().strip()
                        replace = replace_entry.get().strip()
                        if not find:
                            continue
                        if regex_var.get():
                            flags = re.IGNORECASE if case_var.get() else 0
                            new_text = re.sub(find, replace, new_text, flags=flags)
                        elif case_var.get():
                            lower_text = new_text.lower()
                            new_text = lower_text.replace(find.lower(), replace)
                            if new_text != original_text.lower():
                                new_text = original_text.replace(original_text.lower().replace(find.lower(), replace), new_text)
                        else:
                            new_text = new_text.replace(find, replace)
                    if new_text != original_text:
                        if not preview:
                            try:
                                entity.SetCellValue(row, col, new_text)
                            except Exception as e:
                                self.log_message(f"Warning: failed to write table cell on {file_path}: {e}", "error")
                        changes_list.append({
                            "File": file_path, "EntityType": "AcDbTable",
                            "OriginalText": original_text, "NewText": new_text,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
        return changed

    # ---------- DWG processing ----------
    def process_dwg_file(self, acad, file_path, replacements, log, error_log, preview, total_files, processed_files):
        self.log_message(f"Processing: {os.path.basename(file_path)}")
        try:
            for attempt in range(3):
                try:
                    doc = acad.Documents.Open(file_path)
                    pythoncom.PumpWaitingMessages(); time.sleep(1)
                    try:
                        self.apply_stamp_layers(doc, preview)
                    except Exception as e:
                        self.log_message(f"Stamp apply failed: {e}", "error")
                    break
                except Exception as e:
                    if "application is busy" in str(e).lower() and attempt < 2:
                        self.log_message(f"AutoCAD busy, retrying open ({attempt+1}/3)...")
                        time.sleep(2); pythoncom.PumpWaitingMessages(); continue
                    else:
                        raise

            changed = False
            changes_list = []

            for space in (doc.ModelSpace, doc.PaperSpace):
                for entity in space:
                    et = entity.EntityName
                    original_text = None
                    if et == "AcDbText":
                        original_text = entity.TextString
                    elif et == "AcDbMText":
                        try: original_text = entity.Text
                        except Exception:
                            try: original_text = entity.Contents
                            except Exception: original_text = ""
                    elif et.startswith("AcDbDimension"):
                        original_text = entity.TextOverride or ""
                    elif et == "AcDbMLeader" and hasattr(entity, "MTextContent"):
                        original_text = entity.MTextContent

                    if original_text is not None:
                        if self.process_entity_text(entity, et, original_text, replacements, preview, changes_list, file_path):
                            changed = True
                        continue

                    if et == "AcDbBlockReference":
                        if self.process_block_reference(entity, replacements, preview, changes_list, file_path):
                            changed = True
                        continue

                    if et == "AcDbTable":
                        if self.process_table(entity, replacements, preview, changes_list, file_path):
                            changed = True
                        continue

            if changed and not preview:
                doc.Save(); time.sleep(1); pythoncom.PumpWaitingMessages()

            doc.Close(); time.sleep(1); pythoncom.PumpWaitingMessages()

            log.extend(changes_list)
            self.log_message(f"Processed: {os.path.basename(file_path)} - {processed_files + 1}/{total_files}")
            return True

        except Exception as e:
            error_log.append({"File": file_path, "Error": str(e), "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
            self.log_message(f"Error processing {file_path}: {str(e)}", 'error')
            return False

    # ---------- Stamp ----------
    def apply_stamp_layers(self, doc, preview):
        if not self.stamp_apply.get():
            return False
        choice = (self.stamp_issue.get() or "(leave as-is)").strip().upper()
        if choice == "(LEAVE AS-IS)":
            return False
        layer_map = {
            "APPROVAL":     "ISSUE-APPROVAL",
            "AS-BUILT":     "ISSUE-AS-BUILT",
            "BID":          "ISSUE-BID",
            "CONSTRUCTION": "ISSUE-CONSTRUCTION",
            "PRELIM":       "ISSUE-PRELIM",
            "REFERENCE":    "ISSUE-REFERENCE",
        }
        target_layer = layer_map.get(choice)
        if not target_layer:
            return False

        present_layers = []
        try:
            for layer in doc.Layers:
                try: name = layer.Name
                except Exception: continue
                if name.upper().startswith("ISSUE-"):
                    present_layers.append(name)
        except Exception as e:
            self.log_message(f"Stamp: failed to enumerate layers: {e}", "error")
            return False
        if not present_layers:
            self.log_message("Stamp: no ISSUE-* layers found; skipping.")
            return False

        changed = False
        for lname in present_layers:
            try:
                lyr = doc.Layers.Item(lname)
                desired_freeze = (lname.upper() != target_layer.upper())
                if bool(lyr.Freeze) != desired_freeze:
                    if not preview:
                        lyr.Freeze = desired_freeze
                    self.log_message(f"Stamp: {'Froze' if desired_freeze else 'Thawed'} layer {lname}")
                    changed = True
            except Exception as e:
                self.log_message(f"Stamp: could not update layer {lname}: {e}", "error")

        try:
            lyr = doc.Layers.Item(target_layer)
            if lyr.LayerOn is False and not preview:
                lyr.LayerOn = True
        except Exception:
            pass
        return changed

    # ---------- Runner / threading ----------
    def start_process(self):
        if self.process_thread and self.process_thread.is_alive():
            return
        self.cancel_event.clear()
        self.run_btn.config(state=tk.DISABLED)
        self.cancel_btn.config(state=tk.NORMAL)
        self.progress['value'] = 0
        self.process_thread = threading.Thread(target=self.run_process_thread)
        self.process_thread.start()
        self.root.after(100, self.check_queue)

    def cancel_process(self):
        self.cancel_event.set()
        self.log_message("Cancelling process...")

    def check_queue(self):
        try:
            while True:
                msg_type, msg = self.queue.get_nowait()
                if msg_type == 'status':
                    self.log_message(msg)
                elif msg_type == 'error':
                    self.log_message(msg, 'error')
                elif msg_type == 'progress':
                    self.progress['value'] = msg
        except queue.Empty:
            pass
        if self.process_thread and self.process_thread.is_alive():
            self.root.after(100, self.check_queue)
        else:
            self.run_btn.config(state=tk.NORMAL)
            self.cancel_btn.config(state=tk.DISABLED)

    def run_process_thread(self):
        """Main worker; auto-saves + opens the report."""
        # COM must be initialized per-thread when using pywin32 in background threads.
        pythoncom.CoInitialize()
        try:
            path = self.path_entry.get()
            if not path:
                self.log_message("No path selected.", 'error'); return

            replacements = self.validate_replacements(self.replacements)
            if replacements is None:
                return

            # Any non-text ops?
            try:
                clear_count = self.tb_clear_list.size()
            except Exception:
                try:
                    clear_count = len(self.tb_clear_list.get(0, tk.END))
                except Exception:
                    clear_count = 0

            doing_titleblock_ops = (
                self.tb_enable.get() or
                self.tb_shift_down.get() or
                self.tb_shift_up.get() or
                clear_count > 0
            )
            stamp_choice = (self.stamp_issue.get() or "").strip().upper()
            stamp_selected = self.stamp_apply.get() and stamp_choice != "(LEAVE AS-IS)"
            doing_any_ops = doing_titleblock_ops or stamp_selected

            if not replacements and not doing_any_ops:
                self.log_message("No valid replacements entered and no Title Block or Stamp operations selected.", 'error')
                return
            if not replacements and doing_any_ops:
                self.log_message("Proceeding with non-text operations only.")

            preview = self.preview_var.get()

            # File list + report dir
            mode = self.mode_var.get()
            if mode == "directory":
                if not os.path.isdir(path):
                    self.log_message("Invalid directory.", 'error'); return
                dwg_files = []
                if self.recursive_var.get():
                    for root_dir, _, files in os.walk(path):
                        dwg_files.extend(os.path.join(root_dir, f) for f in files if f.lower().endswith('.dwg'))
                else:
                    dwg_files = [os.path.join(path, f) for f in os.listdir(path) if f.lower().endswith('.dwg')]
                try:
                    report_dir = compute_report_dir("directory", path, None)
                except Exception as e:
                    self.log_message(f"Could not resolve report directory: {e}", 'error'); return
            else:
                dwg_files = [p for p in path.split(";") if p]
                if not dwg_files:
                    self.log_message("No DWG files selected.", 'error'); return
                try:
                    report_dir = compute_report_dir("files", None, dwg_files)
                except Exception as e:
                    self.log_message(f"Could not resolve report directory: {e}", 'error'); return

            if not dwg_files:
                self.log_message("No DWG files found.", 'error'); return

            acad = self.get_autocad_app()
            if not acad:
                return

            changes_rows = []; error_rows = []
            total_files = len(dwg_files)
            success_count = 0; fail_count = 0

            for i, file_path in enumerate(dwg_files):
                if self.cancel_event.is_set():
                    self.log_message("Process cancelled."); break
                ok = self.process_dwg_file(acad, file_path, replacements, changes_rows, error_rows, preview, total_files, i)
                if ok:
                    success_count += 1
                else:
                    fail_count += 1
                    self.log_message(f"Skipped {file_path} due to error, continuing...")

            self.log_message(
                f"Process completed. {success_count} succeeded, {fail_count} failed."
                + (" See error log for details." if fail_count else "")
            )

            # Report (always create; then open)
            try:
                os.makedirs(report_dir, exist_ok=True)
                report_name = f"BatchFindandReplace Report {today_mm_dd_yy()}.xlsx"
                report_path = ensure_unique_path(os.path.join(report_dir, report_name))

                df_changes = pd.DataFrame(changes_rows) if changes_rows else pd.DataFrame()
                df_errors  = pd.DataFrame(error_rows)  if error_rows  else pd.DataFrame()
                wb = Workbook()

                # Changes
                ws_changes = wb.active; ws_changes.title = "Changes"
                if not df_changes.empty:
                    headers = list(df_changes.columns)
                    for c, h in enumerate(headers, 1):
                        cell = ws_changes.cell(row=1, column=c, value=h); cell.font = Font(bold=True)
                    for r, row in enumerate(df_changes.itertuples(index=False), 2):
                        for c, val in enumerate(row, 1):
                            cell = ws_changes.cell(row=r, column=c, value=val)
                            if headers[c-1] == "File":
                                cell.hyperlink = f"file:///{str(val).replace('\\', '/')}"
                else:
                    ws_changes.cell(row=1, column=1, value="No changes recorded")

                # Errors
                ws_errors = wb.create_sheet("Errors")
                if not df_errors.empty:
                    headers = list(df_errors.columns)
                    for c, h in enumerate(headers, 1):
                        cell = ws_errors.cell(row=1, column=c, value=h); cell.font = Font(bold=True)
                    for r, row in enumerate(df_errors.itertuples(index=False), 2):
                        for c, val in enumerate(row, 1):
                            ws_errors.cell(row=r, column=c, value=val)
                else:
                    ws_errors.cell(row=1, column=1, value="No errors recorded")

                # Summary
                ws_summary = wb.create_sheet("Summary")
                ws_summary["A1"].font = Font(bold=True); ws_summary["A1"] = "Timestamp"
                ws_summary["B1"].font = Font(bold=True); ws_summary["B1"] = "Total"
                ws_summary["C1"].font = Font(bold=True); ws_summary["C1"] = "Succeeded"
                ws_summary["D1"].font = Font(bold=True); ws_summary["D1"] = "Failed"
                ws_summary["A2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws_summary["B2"] = total_files
                ws_summary["C2"] = success_count
                ws_summary["D2"] = fail_count

                wb.save(report_path)
                self.log_message(f"Report saved to {report_path}")
                open_in_default_app(report_path)
            except Exception as e:
                self.log_message(f"Error saving/opening report: {e}", "error")
        finally:
            # Always uninitialize COM on this thread
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    # ---------- Clear queue UI ----------
    def _ui_add_clear_rev(self):
        choice = (self.tb_clear_choice.get() or "Rev").strip()
        if choice not in ("Rev", "Rev2", "Rev3", "Rev4", "Rev5"):
            return
        if choice not in self.tb_clear_list.get(0, tk.END):
            self.tb_clear_list.insert(tk.END, choice)

    def _ui_remove_selected_clear_rev(self):
        sel = list(self.tb_clear_list.curselection())
        for i in reversed(sel):
            self.tb_clear_list.delete(i)

    # ---------- Suggest/Bug ----------
    def open_suggestion_window(self):
        win = tk.Toplevel(self.root); win.title("Make a Suggestion")
        win.configure(bg=COLOR_BG); win.resizable(False, False)
        tk.Label(win, text="Enter your suggestion:", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=5)
        txt = scrolledtext.ScrolledText(win, wrap="word", bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, height=5)
        txt.pack(padx=10, pady=5, fill="both")
        tk.Label(win, text="Your name (optional):", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=(8, 2))
        name = tk.Entry(win, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, insertbackground='white')
        name.pack(padx=10, pady=(0, 8), fill="x")
        tk.Label(win, text="Your email (optional):", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=(2, 2))
        email = tk.Entry(win, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, insertbackground='white')
        email.pack(padx=10, pady=(0, 8), fill="x")

        def submit():
            s = txt.get("1.0", "end").strip()
            n = name.get().strip() or None
            em = email.get().strip() or None
            if s:
                send_email_single("Suggestion", s, user_name=n, user_email=em)
                self.log_message("Suggestion submitted successfully.")
            else:
                self.log_message("Suggestion submission cancelled: No text entered.", 'error')
            win.destroy()

        tk.Button(win, text="Submit", command=submit, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, relief="flat").pack(pady=5)
        self.log_message("Suggestion window opened.")

    def open_bug_report_window(self):
        win = tk.Toplevel(self.root); win.title("Report a Bug")
        win.configure(bg=COLOR_BG); win.resizable(False, False)
        tk.Label(win, text="Describe the error:", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=5)
        txt = scrolledtext.ScrolledText(win, wrap="word", bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, height=5)
        txt.pack(padx=10, pady=5, fill="both")
        tk.Label(win, text="Your name (optional):", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=(5, 2))
        name = tk.Entry(win, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, insertbackground='white')
        name.pack(padx=10, pady=(0, 8), fill="x")
        tk.Label(win, text="Your email (optional):", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=(2, 2))
        email = tk.Entry(win, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, insertbackground='white')
        email.pack(padx=10, pady=(0, 8), fill="x")

        def submit():
            s = txt.get("1.0", "end").strip()
            n = name.get().strip() or None
            em = email.get().strip() or None
            if s:
                send_email_single("Bug Report", s, user_name=n, user_email=em)
                self.log_message("Bug report submitted successfully.")
            else:
                self.log_message("Bug report submission cancelled: No text entered.", 'error')
            win.destroy()

        tk.Button(win, text="Submit", command=submit, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, relief="flat").pack(pady=5)
        self.log_message("Bug report window opened.")

# ---------- Email ----------
def _escape_html(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def send_email_single(subject: str, body: str, user_name: str | None = None, user_email: str | None = None):
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        submitted_by_display = user_name or "(not provided)"
        submitted_email_display = user_email or "(not provided)"
        html = f"""
        <div style="background:#121212;color:#FFFFFF;font-family:Arial,sans-serif;padding:20px;">
          <div style="text-align:center;margin-bottom:15px;">
            <img src="cid:root3logo" alt="Root3Power Logo" style="width:80px;height:80px;"/>
            <h2 style="color:#1E90FF;margin:10px 0 0;">Root3Power LLC</h2>
            <p style="margin:0;font-size:14px;color:#4682B4;">{_escape_html(subject)}</p>
          </div>
          <div style="background:#1E1E1E;padding:15px;border-radius:8px;">
            <p><strong>Submitted By (name):</strong> {_escape_html(submitted_by_display)}</p>
            <p><strong>Submitted By (email):</strong> {_escape_html(submitted_email_display)}</p>
            <p><strong>Submitted At:</strong> {timestamp}</p>
            <p><strong>Message:</strong></p>
            <div style="background:#2C2C2C;padding:10px;border-radius:6px;color:#FFFFFF;white-space:pre-wrap;">
              {_escape_html(body)}
            </div>
          </div>
          <div style="margin-top:12px;color:#BBBBBB;font-size:12px;">
            App Version: {VERSION}
          </div>
        </div>
        """
        msg = MIMEMultipart("related")
        msg["Subject"] = subject
        msg["From"] = formataddr(("Batch Find & Replace", SENDER_EMAIL))
        msg["To"] = RECEIVER_EMAIL
        if user_email:
            msg["Reply-To"] = user_email

        alt = MIMEMultipart("alternative"); msg.attach(alt)
        alt.attach(MIMEText(html, "html"))

        try:
            with open(resource_path(LOGO_FILE), "rb") as f:
                logo = MIMEImage(f.read())
                logo.add_header("Content-ID", "<root3logo>")
                logo.add_header("Content-Disposition", "inline", filename=LOGO_FILE)
                msg.attach(logo)
        except Exception:
            pass

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, APP_PASSWORD)
            server.sendmail(SENDER_EMAIL, [RECEIVER_EMAIL], msg.as_string())
    except Exception as e:
        logger.error(f"Error sending email: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = AutoCADReplacerGUI(root)
    root.mainloop()
